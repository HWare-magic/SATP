import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors

wb =openpyxl.Workbook()


#font
ws =wb.active
ws.title = 'font'
#default 11pt,cabri
italic24font =Font(size=24,italic=True)
ws['B3'].font = italic24font
ws['B3']='sjhdfui'
boldRedfont =Font(name='Times New Roman',bold=True,color=colors.BLUE)
ws['B3'].font= boldRedfont

#formulas
ws =wb.create_sheet('formula')
ws['A1']=200
ws['A2']=200
ws['A3']= '=SUM(A1:A2)'
#setting row heigh and columu
ws =wb.create_sheet('dime')
ws['A1'] = 'tall row'
ws.row_dimensions[1].height =70
ws['B2'] ='width col'
ws.column_dimensions['B'].width =20

#merge cells

ws =wb.create_sheet('merge')
ws.merge_cells('A1:D3')
ws['A1'] ='merge cell'

#unmergr cell
ws= wb.copy_worksheet(wb.get_sheet_by_name('merge'))
ws.title = 'unmerge'
ws.unmerge_cells('A1:D3')

wb.save('fomate.xlsx')