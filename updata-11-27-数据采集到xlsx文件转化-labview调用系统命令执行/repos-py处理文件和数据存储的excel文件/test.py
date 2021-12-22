import openpyxl
import csv
# turn txt file to csv file  
csvFile = open("data.csv",'w',newline='',encoding='utf-8')
writer = csv.writer(csvFile)
csvRow=[]

f = open(r'C:\Users\86136\Desktop\1msVt of temp.txt','r',encoding='GB2312')
for line in f:
    csvRow=line.split()
    writer.writerow(csvRow)
f.close()
csvFile.close()
# data process
wb = openpyxl.load_workbook("data.csv")
sheet = wb.active
wb.creat_sheet(index=1,title='res_value')
ws = wb['res_value']
for num in range (1,11):
    for row in range (1+num*1070,31+num*1070):
        for col in range (2,33):
            ws.cell(column=col,row=row-num*1040,value=sheet.cell(column=col,row=row).value + 5)
            #ws.cell(column=col,row=row-num*1040,value=(5-sheet.cell(column=col,row=row).value)*5/sheet.cell(column=col,row=row).value)
wb.creat_sheet(index=2,title='return_sheet')
ws1 = wb['return_sheet']
for cle in range (1,3):
    for plus in range (1,5):
        for col in range (2,33):
            ws1.cell(column=col,row=cle,value=(ws1.cell(column=col,row=cle).value + ws.cell(column=col,row=25+plus+30*cle).value))

wb.save("data1.csv")